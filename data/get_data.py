from openpyxl import load_workbook
import numpy as np

file_path = 'data/TCGA_6_Cancer_Type_Mutation_List.xlsx'

def get_data(file_path):
    def is_number(s):
        try:
            float(s)
            return True
        except ValueError:
            return False

    def transferRef(Ref):
        temp = ''
        dist_Ref = {'A': '1', 'C': '2', 'G': '3', 'T': '4'}
        if Ref in dist_Ref:
            return dist_Ref[Ref]
        elif Ref == "-":
            return '0'
        else:
            for letter in Ref:
                if temp == '':
                    temp = temp + str(dist_Ref[letter])
                else:
                    temp = temp + ',' + str(dist_Ref[letter])
            return temp


    workbook = load_workbook(file_path, use_iterators=True, read_only=True)
    first_sheet = workbook.get_sheet_names()[0]
    worksheet = workbook.get_sheet_by_name(first_sheet)
    FirstRow = True
    dict_CancerType = {}
    dict_TumorID = {}
    dict_GeneName = {}
    dict_Chromosome = {}
    dict_VariantType = {}
    data = []

    for row in worksheet.iter_rows():
        Row = []
        if FirstRow:
            FirstRow = False
            continue

        # CANCER_TYPE
        cancerType = str(row[0].value)
        if cancerType in dict_CancerType:
            Row.append(dict_CancerType[cancerType])
        else:
            nCancer = len(dict_CancerType)
            dict_CancerType[cancerType] = nCancer
            Row.append(dict_CancerType[cancerType])

        # TUMOR_SAMPLE_ID
        tumorID = str(row[1].value)
        if tumorID in dict_TumorID:
            Row.append(dict_TumorID[tumorID])
        else:
            nTumorId = len(dict_TumorID)
            dict_TumorID[tumorID] = nTumorId + 1
            Row.append(dict_TumorID[tumorID])


        # GENE_NAME
        geneName = str(row[2].value)
        if geneName in dict_GeneName:
            Row.append(dict_GeneName[geneName])
        else:
            nGene = len(dict_GeneName)
            dict_GeneName[geneName] = nGene + 1
            Row.append(dict_GeneName[geneName])


        #CHROMOSOME

        Chromosome = str(row[3].value)
        if is_number(Chromosome):
            Row.append(float(Chromosome))
        else:
            if Chromosome in dict_Chromosome:
                Row.append(dict_Chromosome[Chromosome])
            else:
                nChromo = len(dict_Chromosome)
                dict_Chromosome[Chromosome] = (nChromo + 1)*(-1)
                Row.append(dict_Chromosome[Chromosome])

        #START POSITION

        startPos = float(row[4].value)
        Row.append(startPos)

        #END_POSITION
        endPos = float(row[5].value)
        Row.append(endPos)

        #VARIANT TYPE
        variantType = str(row[6].value)
        if variantType in dict_VariantType:    #TUMOR_ALLELE

            Row.append(dict_VariantType[variantType])
        else:
            nVariant = len(dict_VariantType)
            dict_VariantType[variantType] = nVariant + 1
            Row.append(dict_VariantType[variantType])

        #REFERENCE_ALLELE
        referAllele = row[7].value
        Row.append(transferRef(referAllele))

        #TUMOR_ALLELE
        tumorAlle = row[8].value
        Row.append(transferRef(tumorAlle))

        data.append(Row)

    return np.array(data)

data = get_data(file_path)
np.save('data.npy', data)

